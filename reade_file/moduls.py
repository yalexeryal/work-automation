import csv
import os
import datetime
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


def reade_file_xlsx(file, num=0):
    """
    Reads xlsx file
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb.active
    return sheet


# def create_file_xls(file):
#     today = datetime.datetime.today().date()
#     df = pd.read_excel(file)
#     df_rez = df[['БЮ', 'Продукт', 'Модуль', 'Название задания', 'ID задания', 'Ссылка на работу в админке',
#                  'Ссылка на работу в ЛК эксперта', 'Студент', 'Дедлайн', 'Дедлайн', 'Отправлена', 'Проверющий',
#                  'Возможные проверяющие', 'Дней на проверке']]
#
#
#     df_rez_sorted = df_rez.sort_values('Отправлена')
#
#     result_file = f"result_file/Непроверенные ДЗ {today}.xlsx"
#     df_rez_sorted.to_excel(result_file, index=False)
#     return f"Создан файл {os.path.basename(result_file)}. В папке по адресу: {os.path.abspath(result_file)}"


def create_file_xls(file):
    today = datetime.datetime.today().date()
    df = pd.read_excel(file)
    df_rez = df[['БЮ', 'Продукт', 'Модуль', 'Название задания', 'ID задания', 'Ссылка на работу в админке',
                 'Ссылка на работу в ЛК эксперта', 'Студент', 'Дедлайн', 'Дедлайн', 'Отправлена', 'Проверющий',
                 'Возможные проверяющие', 'Дней на проверке']]

    df_rez_sorted = df_rez.sort_values('Отправлена')

    result_file = f"result_file/Непроверенные ДЗ {today}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF")

    for col_num, column_title in enumerate(df_rez_sorted.columns, 1):
        cell = sheet.cell(row=1, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    for row_num, row_data in enumerate(df_rez_sorted.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(result_file)
    return f"Создан файл {os.path.basename(result_file)}. В папке по адресу: {os.path.abspath(result_file)}"

def reade_file_csv(file: csv):
    with open(file, encoding="utf-8") as f:
        reader = csv.reader(f)
        synopsis_list = list(reader)
        return synopsis_list


def delete_startup_file():
    """
    Delete the original data file from the work_file folder.
    """
    folder_path = os.path.dirname("work_file/")
    file_list = os.listdir(folder_path)
    for file_name in file_list:
        file_path = os.path.join(folder_path, file_name)
        if os.path.isfile(file_path):
            os.remove(file_path)


if __name__ == '__main__':
    dir_name = "../result_file/"
    dirs = os.path.dirname("../work_file/")
    file = f"{dirs}/{os.listdir(path=dirs)[0]}"
    sheet = reade_file_xlsx(file)
    create_file_xls(sheet, dir_name)

