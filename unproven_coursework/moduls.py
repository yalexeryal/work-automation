import os
import datetime

import pandas as pd

from db.moduls import reade_db_file
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def checker_inspector(checker, inspectors) -> str:
    """
    Bringing out the checker.
    """
    if checker is not None:
        return checker
    elif len(inspectors) == 1 and inspectors is not None:
        return str(inspectors[0])


def date_filter(overdue: datetime, tomorrow_start: datetime, submitted: datetime, row: tuple):
    """
    Enter dates and a row from the works list, give the list index to add and a row to output.

    """
    overdue_date = overdue
    tomorrow_start_date = tomorrow_start
    modul = row[2]
    session = row[3]
    session_link = row[12]
    student = row[8]

    if submitted <= overdue_date:
        return 1, f"{modul}    {session}    {session_link}    {student}    {submitted}"

    elif overdue_date < submitted < tomorrow_start_date:
        return 2, f"{modul}    {session}    {session_link}    {student}"

    elif tomorrow_start_date <= submitted:
        return 3, f"{modul}    {session}    {session_link}    {student}"


def write_expert_dip_file(experts_dict: dict, rez_file_expert, soft_expert):
    """
    Write data from the "experts_dict" dictionary into the file
    """
    with open(rez_file_expert, 'w', encoding='utf-8') as f:
        for ex_d in experts_dict.items():
            if ex_d[0] in soft_expert:
                if len(ex_d[1][1]) > 1:
                    name_exp = f"\n\n {ex_d[0]}  \nПривет!\n"
                    f.write(name_exp)
                    for hw in ex_d[1][1]:
                        rez = f"{hw}\n"
                        f.write(rez)
            else:
                name_exp = f"\n\n {ex_d[0]}  \nПривет!\n"
                f.write(name_exp)
                if len(ex_d[1][1]) > 1:
                    for hw in ex_d[1][1]:
                        rez = f"{hw}\n"
                        f.write(rez)
                if len(ex_d[1][2]) > 1:
                    for hw in ex_d[1][2]:
                        rez = f"{hw}\n"
                        f.write(rez)
                if len(ex_d[1][3]) > 1:
                    for hw in ex_d[1][3]:
                        rez = f"{hw}\n"
                        f.write(rez)
    return f"Создан файл {os.path.basename(rez_file_expert)}. " \
           f"В папке по адресу: {os.path.abspath(rez_file_expert)}"


def sorted_dict(unsorted_dictionary: dict):
    sorted_tuple = sorted(unsorted_dictionary.items(), key=lambda x: x[0])
    return dict(sorted_tuple)


def create_dip_dict(sheet, files):
    """
    Creating Dictionaries
    experts_dict - unverified tasks with assigned experts
    """

    experts_dict = {}
    checker_none = []
    diploma_blocks = reade_db_file(files)
    today = datetime.date.today()
    overdue = today - datetime.timedelta(days=8)
    tomorrow_start = today - datetime.timedelta(days=6)
    tomorrow_finish = today - datetime.timedelta(days=4)
    new_list_expert = ['Привет!',
                       ['Есть просроченные курсовые, необходимо проверить в ближайшее время:\n'],
                       ['\nСегодня до конца дня необходимо проверить курсовые:\n'],
                       ['\nНа всякий случай - до завтра до конца дня необходимо проверить курсовые:\n'],
                       ]


    for row in sheet.iter_rows(min_row=2, values_only=True):
        # if row[0] == 'Программирование':
        submitted = datetime.datetime.strptime(row[9], "%Y-%m-%d").date()
        moduls = row[1].upper()
        checker = row[10]
        outcome = date_filter(overdue, tomorrow_start, submitted, row)
        ind, rez = outcome[0], outcome[1]
        session = row[5]
        session_link = row[12]
        student = row[8]

        if moduls in diploma_blocks:
            continue
        else:
            if checker is None:
                checker_none.append(f"{row[2]}  {session}  {session_link}  {student}  {submitted}")
            elif submitted <= tomorrow_finish:
                outcome = date_filter(overdue, tomorrow_start, submitted, row)
                ind, rez = outcome[0], outcome[1]

                if checker not in experts_dict:
                    experts_dict[checker] = new_list_expert.copy()
                    experts_dict[checker][ind] = experts_dict[checker][ind] + [rez]

                else:
                    experts_dict[checker][ind] = experts_dict[checker][ind] + [rez]

    for i in checker_none:
        print(i)

    return sorted_dict(experts_dict)

def create_file_dip_xls(file):
    today = datetime.datetime.today().date()
    df = pd.read_excel(file)
    df_rez = df[['Модуль', 'Название задания', 'Ссылка на работу в админке',
                 'Ссылка на работу в ЛК эксперта', 'Студент', 'Отправлена', 'Проверющий',
                 'Возможные проверяющие']]

    df_rez_sorted = df_rez.sort_values('Модуль')

    result_file = f"result_file/Непроверенные дипломы {today}.xlsx"
    df_rez_sorted.to_excel(result_file, index=False)
    workbook = Workbook()
    sheet = workbook.active
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF")

    for col_num, column_title in enumerate(df_rez_sorted.columns, 1):
        cell = sheet.cell(row=1, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font

    for row_num, row_data in enumerate(df_rez_sorted.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(result_file)
    return f"Создан файл {os.path.basename(result_file)}. В папке по адресу: {os.path.abspath(result_file)}"

# if __name__ == '__main__':
#     create_dict()
